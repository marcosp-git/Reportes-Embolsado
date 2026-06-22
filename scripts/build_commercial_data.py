#!/usr/bin/env python3
"""Build the local browser dataset for the Embolsado commercial map.

This writes public/commercial-data.js for local preview. The file is ignored by
git because it contains customer-level commercial data.
"""

from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORK = ROOT / "Archivos de Trabajo"
AUDIT = ROOT / "outputs" / "commercial-audit"
OUT = ROOT / "public" / "commercial-data.js"


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def norm_id(value: Any) -> str:
    text = clean(value).upper()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def seller_key(value: Any) -> str:
    text = norm_id(value)
    if text.isdigit() and len(text) < 3:
        return text.zfill(3)
    return text


def parse_number(value: Any) -> float:
    text = clean(value)
    if not text:
        return 0.0
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def read_sheet(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        header = [clean(value) for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({header[i]: row[i] if i < len(row) else None for i in range(len(header)) if header[i]})
        return rows
    finally:
        wb.close()


def top_items(counter: Counter[str], limit: int = 6) -> list[dict[str, Any]]:
    return [{"name": name, "um": round(value, 2)} for name, value in counter.most_common(limit) if value]


def seller_names(workbook_path: Path) -> dict[str, str]:
    counters: defaultdict[str, Counter[str]] = defaultdict(Counter)
    if not workbook_path.exists():
        return {}

    for sheet_name in ("SM", "VEN MM I"):
        try:
            rows = read_sheet(workbook_path, sheet_name)
        except KeyError:
            continue
        for row in rows:
            code = seller_key(row.get("VENDEDOR") or row.get("V"))
            name = clean(row.get("VENDEDOR_"))
            if code and name:
                counters[code][name.upper()] += 1

    return {code: counter.most_common(1)[0][0] for code, counter in counters.items() if counter}


def seller_display(code: str, names: dict[str, str]) -> str:
    normalized = seller_key(code)
    name = names.get(normalized)
    if name and name != normalized:
        return f"{name} ({normalized})"
    return normalized or clean(code)


def main() -> None:
    clients_path = AUDIT / "clients_normalized.csv"
    coverage_path = AUDIT / "fact_source_coverage.csv"
    zone_path = AUDIT / "zone_summary.csv"
    conflict_path = AUDIT / "status_conflicts.csv"
    missing_path = AUDIT / "clients_without_coordinates.csv"
    volume_path = WORK / "VENDEDORES PARTICULARES" / "Volumen por Cliente" / "Volumen por cliente.xlsx"
    clients_total_path = WORK / "VENDEDORES PARTICULARES" / "CLIENTES TOTALES.xlsx"

    if not clients_path.exists():
        raise SystemExit("Run scripts/audit_commercial_data.py first.")

    client_rows = read_csv(clients_path)
    clients_by_id = {norm_id(row["client_id"]): row for row in client_rows if norm_id(row.get("client_id"))}
    seller_name_by_code = seller_names(clients_total_path)

    total_um: defaultdict[str, float] = defaultdict(float)
    family_um: defaultdict[str, Counter[str]] = defaultdict(Counter)
    product_um: defaultdict[str, Counter[str]] = defaultdict(Counter)

    if volume_path.exists():
        for row in read_sheet(volume_path, "C TOT"):
            client_id = norm_id(row.get("Cliente"))
            if not client_id:
                continue
            total_um[client_id] = max(total_um[client_id], parse_number(row.get("UM")))

        for row in read_sheet(volume_path, "SM"):
            client_id = norm_id(row.get("Cliente"))
            if not client_id:
                continue
            product = clean(row.get("Producto"))
            if not product:
                continue
            um = parse_number(row.get("UM"))
            product_um[client_id][product] += um

        for row in read_sheet(volume_path, "ART"):
            client_id = norm_id(row.get("Cliente"))
            if not client_id:
                continue
            um = parse_number(row.get("UM"))
            family = clean(row.get("FAMILIA")) or "ART"
            family_um[client_id][family] += um

        for sheet, family_name, id_key in (
            ("000", "000", "Clientes"),
            ("0000", "0000", "CLIENTES"),
            ("HE", "H ESP", "Cliente"),
            ("T", "TAPERA", "Cliente"),
            ("SA", "SALVADO", "Cliente"),
            ("SE", "SEMOLIN", "Cliente"),
        ):
            try:
                rows = read_sheet(volume_path, sheet)
            except KeyError:
                continue
            for row in rows:
                client_id = norm_id(row.get(id_key))
                if not client_id:
                    continue
                family_um[client_id][family_name] += parse_number(row.get("UM"))

    map_clients = []
    for client_id, row in clients_by_id.items():
        if not row.get("lat") or not row.get("lon"):
            continue
        um = total_um.get(client_id, 0.0)
        map_clients.append(
            {
                "id": client_id,
                "name": row.get("name") or "",
                "seller": row.get("seller") or "",
                "sellerDisplay": seller_display(row.get("seller") or "", seller_name_by_code),
                "status": row.get("status") or "SIN ESTADO",
                "zoneId": row.get("zone_id") or "interior",
                "zoneName": row.get("zone_name") or "Interior",
                "lat": float(row["lat"]),
                "lon": float(row["lon"]),
                "totalUm": round(um, 2),
                "families": top_items(family_um[client_id], 5),
                "products": top_items(product_um[client_id], 5),
                "sources": (row.get("sources") or "").split("|") if row.get("sources") else [],
            }
        )

    map_clients.sort(key=lambda item: item["totalUm"], reverse=True)

    zone_summary = read_csv(zone_path) if zone_path.exists() else []
    coverage = read_csv(coverage_path) if coverage_path.exists() else []

    status_counts = Counter(client["status"] for client in map_clients)
    zone_counts = Counter(client["zoneName"] for client in map_clients)
    volume_by_zone: defaultdict[str, float] = defaultdict(float)
    for client in map_clients:
        volume_by_zone[client["zoneName"]] += client["totalUm"]

    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "summary": {
            "clients": len(client_rows),
            "mapClients": len(map_clients),
            "withVolume": sum(1 for client in map_clients if client["totalUm"] > 0),
            "missingCoordinates": max(0, sum(1 for row in client_rows if not row.get("lat"))),
            "statusConflicts": max(0, len(read_csv(conflict_path)) if conflict_path.exists() else 0),
            "withoutCoordinatesRows": max(0, len(read_csv(missing_path)) if missing_path.exists() else 0),
        },
        "statusCounts": dict(sorted(status_counts.items())),
        "zoneCounts": dict(sorted(zone_counts.items())),
        "volumeByZone": {zone: round(value, 2) for zone, value in sorted(volume_by_zone.items())},
        "zoneSummary": zone_summary,
        "coverage": coverage,
        "clients": map_clients,
    }

    OUT.write_text("window.EMBOLSADO_COMMERCIAL_DATA = " + json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + ";\n", encoding="utf-8")
    print(f"Wrote {OUT} with {len(map_clients)} mapped clients")


if __name__ == "__main__":
    main()
