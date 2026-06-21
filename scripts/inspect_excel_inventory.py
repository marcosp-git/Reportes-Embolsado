#!/usr/bin/env python3
"""Inventory the Embolsado Excel workbooks for dashboard modeling."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORK = ROOT / "Archivos de Trabajo" / "Informes Iñaqui"
OUT = ROOT / "outputs" / "commercial-audit" / "excel_inventory.json"

KEYWORDS = (
    "OBJET",
    "CUMPL",
    "AVANCE",
    "PROY",
    "JEFE",
    "VENDED",
    "CORRED",
    "CLIENT",
    "TON",
    "KG",
    "UM",
    "REAL",
    "PRESUP",
    "%",
)


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def sheet_sample(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True):
        values = [clean(value) for value in row[:24]]
        if any(values):
            rows.append(values)
    return rows


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    inventory = []

    for path in sorted(WORK.glob("*.xlsx")):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            file_info = {"file": path.name, "sheets": []}
            for ws in wb.worksheets:
                sample = sheet_sample(ws)
                text = " ".join(" ".join(row) for row in sample).upper()
                score = sum(1 for keyword in KEYWORDS if keyword in text)
                file_info["sheets"].append(
                    {
                        "sheet": ws.title,
                        "rows": ws.max_row,
                        "cols": ws.max_column,
                        "score": score,
                        "sample": sample[:6],
                    }
                )
            inventory.append(file_info)
        finally:
            wb.close()

    OUT.write_text(json.dumps(inventory, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
