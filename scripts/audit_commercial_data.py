#!/usr/bin/env python3
"""Build a local commercial data audit from the Embolsado working files.

The generated outputs can contain sensitive customer data and are written under
outputs/, which is intentionally git-ignored.
"""

from __future__ import annotations

import csv
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORK = ROOT / "Archivos de Trabajo"
OUT = ROOT / "outputs" / "commercial-audit"
REPORT = OUT / "diagnostic_report.md"

FACT_SPECS = [
    ("ACTIVIDAD JUNIO / BASE", "cliente categoria", WORK / "Informes Iñaqui" / "ACTIVIDAD JUNIO 1.xlsx", "BASE", 2, 0),
    ("ACTIVIDAD JUNIO / ACTIVIDAD DE CLIENTES", "cliente-semana estado", WORK / "Informes Iñaqui" / "ACTIVIDAD JUNIO 1.xlsx", "ACTIVIDAD DE CLIENTES", 3, 2),
    ("Clientes Nuevos / Nuevos", "cliente nuevo-producto", WORK / "Informes Iñaqui" / "Clientes Nuevos y Recuperados 2.xlsx", "Nuevos", 2, 0),
    ("Clientes Por Tipo / Total", "cliente-producto familias", WORK / "Informes Iñaqui" / "Clientes Por Tipo de Producto.xlsx", "Total", 5, 0),
    ("Clientes Por Tipo / Datos", "cliente-producto raw", WORK / "Informes Iñaqui" / "Clientes Por Tipo de Producto.xlsx", "Datos", 2, 1),
    ("Ranking Clientes / Tabla No Formulada", "cliente-mes ranking", WORK / "Informes Iñaqui" / "Ránking de Clientes 1.xlsx", "Tabla No Formulada", 3, 0),
    ("PP2SJ / GEN", "cliente-producto importe", WORK / "Informes Iñaqui" / "PP2SJ 1.xlsx", "GEN", 2, 4),
    ("Volumen por Cliente / VOLUMEN POR CLIENTE", "cliente geocodificado", WORK / "VENDEDORES PARTICULARES" / "Volumen por Cliente" / "Volumen por cliente.xlsx", "VOLUMEN POR CLIENTE", 2, 1),
    ("Volumen por Cliente / C TOT", "cliente volumen total", WORK / "VENDEDORES PARTICULARES" / "Volumen por Cliente" / "Volumen por cliente.xlsx", "C TOT", 2, 1),
    ("Volumen por Cliente / SM", "cliente-producto volumen", WORK / "VENDEDORES PARTICULARES" / "Volumen por Cliente" / "Volumen por cliente.xlsx", "SM", 2, 1),
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return unicodedata.normalize("NFC", text)


def norm_id(value: Any) -> str:
    text = clean_text(value).upper()
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text


def norm_header(value: Any) -> str:
    text = clean_text(value).upper()
    text = text.replace("Á", "A").replace("É", "E").replace("Í", "I")
    text = text.replace("Ó", "O").replace("Ú", "U").replace("Ñ", "N")
    return re.sub(r"\s+", " ", text)


def parse_number(value: Any) -> float | None:
    text = clean_text(value)
    if not text:
        return None
    text = text.replace(".", "").replace(",", ".") if "," in text else text
    try:
        return float(text)
    except ValueError:
        return None


def parse_lat_lon(lat: Any = None, lon: Any = None, coord: Any = None) -> tuple[float, float] | None:
    if coord not in (None, ""):
        parts = re.split(r"[,;]", clean_text(coord))
        if len(parts) >= 2:
            parsed = parse_lat_lon(parts[0], parts[1])
            if parsed:
                return parsed

    lat_num = parse_number(lat)
    lon_num = parse_number(lon)
    if lat_num is None or lon_num is None:
        return None
    if -56 <= lat_num <= -20 and -75 <= lon_num <= -50:
        return (lat_num, lon_num)
    return None


def read_csv_dicts(path: Path) -> list[dict[str, str]]:
    raw = path.read_bytes()
    text = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            pass
    if text is None:
        text = raw.decode("latin-1", errors="replace")

    sample = "\n".join(text.splitlines()[:5])
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    return list(csv.DictReader(text.splitlines(), delimiter=delimiter))


@dataclass
class ClientRecord:
    client_id: str
    names: Counter[str] = field(default_factory=Counter)
    sellers: Counter[str] = field(default_factory=Counter)
    statuses: defaultdict[str, set[str]] = field(default_factory=lambda: defaultdict(set))
    categories: Counter[str] = field(default_factory=Counter)
    addresses: Counter[str] = field(default_factory=Counter)
    coords: list[tuple[float, float, str]] = field(default_factory=list)
    sources: set[str] = field(default_factory=set)
    zone_id: str = "interior"
    zone_name: str = "Interior"

    def add(
        self,
        source: str,
        name: Any = None,
        seller: Any = None,
        status: Any = None,
        category: Any = None,
        address: Any = None,
        coord: tuple[float, float] | None = None,
    ) -> None:
        self.sources.add(source)
        if clean_text(name):
            self.names[clean_text(name)] += 1
        if clean_text(seller):
            self.sellers[clean_text(seller)] += 1
        if clean_text(status):
            self.statuses[source].add(clean_text(status).upper())
        if clean_text(category):
            self.categories[clean_text(category)] += 1
        if clean_text(address):
            self.addresses[clean_text(address)] += 1
        if coord:
            self.coords.append((coord[0], coord[1], source))

    @property
    def name(self) -> str:
        return self.names.most_common(1)[0][0] if self.names else ""

    @property
    def seller(self) -> str:
        return self.sellers.most_common(1)[0][0] if self.sellers else ""

    @property
    def category(self) -> str:
        return self.categories.most_common(1)[0][0] if self.categories else ""

    @property
    def address(self) -> str:
        return self.addresses.most_common(1)[0][0] if self.addresses else ""

    @property
    def coord(self) -> tuple[float, float] | None:
        if not self.coords:
            return None
        priority = {
            "volumen_por_cliente": 0,
            "clientes_totales": 1,
            "clientes_totales_sheet": 1,
            "csv_activos_inactivos": 2,
            "umap": 3,
        }
        return sorted(self.coords, key=lambda x: priority.get(x[2], 9))[0][:2]

    @property
    def status_values(self) -> set[str]:
        values: set[str] = set()
        for source_values in self.statuses.values():
            values.update(source_values)
        return values

    @property
    def status(self) -> str:
        values = self.status_values
        if "A" in values and "I" in values:
            return "CONFLICTO A/I"
        if "A" in values:
            return "A"
        if "I" in values:
            return "I"
        return ""


def get_client(clients: dict[str, ClientRecord], client_id: Any) -> ClientRecord | None:
    cid = norm_id(client_id)
    if not cid:
        return None
    if cid not in clients:
        clients[cid] = ClientRecord(cid)
    return clients[cid]


def add_csv_clients(clients: dict[str, ClientRecord]) -> None:
    specs = [
        (WORK / "VENDEDORES PARTICULARES" / "UMAP" / "CLIENTES ACTIVOS.csv", "A", "umap"),
        (WORK / "VENDEDORES PARTICULARES" / "UMAP" / "CLIENTES INACTIVOS.csv", "I", "umap"),
    ]
    for path, status, source in specs:
        if not path.exists():
            continue
        for row in read_csv_dicts(path):
            client = get_client(clients, row.get("NROCTA"))
            if not client:
                continue
            client.add(
                source=source,
                name=row.get("CLIENTE"),
                seller=row.get("VENDEDOR"),
                status=status,
                address=row.get("DOMICILIO"),
                coord=parse_lat_lon(row.get("LATITUD"), row.get("LONGITUD")),
            )

    for folder, status in (("ACTIVOS", "A"), ("INACTIVOS", "I")):
        base = WORK / "VENDEDORES PARTICULARES" / folder
        if not base.exists():
            continue
        for path in sorted(base.glob("*.csv")):
            for row in read_csv_dicts(path):
                lat = row.get("LATITUD") or row.get("LATITTUD")
                client = get_client(clients, row.get("NROCTA"))
                if not client:
                    continue
                client.add(
                    source="csv_activos_inactivos",
                    name=row.get("CLIENTE") or row.get("RAZÓN SOCIAL"),
                    seller=row.get("VENDEDOR"),
                    status=status,
                    address=row.get("DOMICILIO"),
                    coord=parse_lat_lon(lat, row.get("LONGITUD")),
                )


def rows_from_sheet(path: Path, sheet_name: str) -> Iterable[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        header = [norm_header(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            yield {header[i]: row[i] if i < len(row) else None for i in range(len(header)) if header[i]}
    finally:
        wb.close()


def add_clientes_totales(clients: dict[str, ClientRecord]) -> None:
    path = WORK / "VENDEDORES PARTICULARES" / "CLIENTES TOTALES.xlsx"
    if not path.exists():
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            header = [norm_header(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            if "NROCTA" not in header:
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = {header[i]: row[i] if i < len(row) else None for i in range(len(header)) if header[i]}
                client = get_client(clients, data.get("NROCTA"))
                if not client:
                    continue
                coord = parse_lat_lon(coord=data.get("COORDENADAS"))
                client.add(
                    source="clientes_totales_sheet" if ws.title not in ("VEN MM", "VEN MM I") else "clientes_totales",
                    name=data.get("CLIENTE"),
                    seller=data.get("VENDEDOR") or data.get("V"),
                    status=data.get("ESTADO"),
                    address=data.get("DOMICILIO"),
                    coord=coord,
                )
    finally:
        wb.close()


def add_volumen_clients(clients: dict[str, ClientRecord]) -> None:
    path = WORK / "VENDEDORES PARTICULARES" / "Volumen por Cliente" / "Volumen por cliente.xlsx"
    if not path.exists():
        return
    for sheet, id_key, name_key in (
        ("VOLUMEN POR CLIENTE", "NROCTA", "CLIENTE"),
        ("C TOT", "CLIENTE", "RAZON SOCIAL"),
        ("000", "CLIENTES", "RAZON SOCIAL"),
        ("0000", "CLIENTES", "RAZON SOCIAL"),
        ("HE", "CLIENTE", "RAZON SOCIAL"),
        ("T", "CLIENTE", "RAZON SOCIAL"),
        ("SA", "CLIENTE", "RAZON SOCIAL"),
        ("SE", "CLIENTE", "RAZON SOCIAL"),
    ):
        try:
            rows = rows_from_sheet(path, sheet)
        except KeyError:
            continue
        for data in rows:
            client = get_client(clients, data.get(id_key))
            if not client:
                continue
            coord = parse_lat_lon(data.get("LATITUD") or data.get("LATITUD"), data.get("LONGITUD"))
            client.add(
                source="volumen_por_cliente",
                name=data.get(name_key),
                seller=data.get("VENDEDOR"),
                coord=coord,
            )


def parse_js_assignment(path: Path, global_name: str) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8")
    match = re.search(rf"window\.{re.escape(global_name)}\s*=\s*(\{{.*\}});?\s*$", text, re.S)
    if not match:
        raise RuntimeError(f"Could not parse {path}")
    return json.loads(match.group(1))


def load_map_zones() -> tuple[list[dict[str, Any]], list[str]]:
    data = parse_js_assignment(ROOT / "public" / "data.js", "EMBOLSADO_MAP_DATA")
    zones = list(data["zones"])
    caba_path = ROOT / "public" / "caba-zones.js"
    if caba_path.exists():
        caba_data = parse_js_assignment(caba_path, "EMBOLSADO_CABA_ZONES")
        zones.extend(caba_data.get("zones", []))
    priority = data.get("territoryAssignment", {}).get("priority") or [z["id"] for z in zones]
    return zones, priority


def point_in_ring(lat: float, lon: float, ring: list[list[float]]) -> bool:
    inside = False
    n = len(ring)
    if n < 3:
        return False
    x, y = lon, lat
    for i in range(n):
        lat1, lon1 = ring[i]
        lat2, lon2 = ring[(i + 1) % n]
        x1, y1 = lon1, lat1
        x2, y2 = lon2, lat2
        if ((y1 > y) != (y2 > y)) and (x < (x2 - x1) * (y - y1) / ((y2 - y1) or 1e-12) + x1):
            inside = not inside
    return inside


def iter_rings(shape: Any) -> Iterable[list[list[float]]]:
    if not isinstance(shape, list) or not shape:
        return
    if isinstance(shape[0], list) and len(shape[0]) >= 2 and all(isinstance(v, (int, float)) for v in shape[0][:2]):
        yield shape
        return
    for item in shape:
        yield from iter_rings(item)


def zone_for_point(lat: float, lon: float, zones: list[dict[str, Any]], priority: list[str]) -> tuple[str, str]:
    by_id = {zone["id"]: zone for zone in zones}
    for zone_id in priority:
        zone = by_id.get(zone_id)
        if not zone:
            continue
        for ring in iter_rings(zone.get("coordinates")):
            if point_in_ring(lat, lon, ring):
                return zone["id"], zone["name"]
    return "interior", "Interior"


def assign_zones(clients: dict[str, ClientRecord]) -> None:
    zones, priority = load_map_zones()
    for client in clients.values():
        coord = client.coord
        if not coord:
            continue
        client.zone_id, client.zone_name = zone_for_point(coord[0], coord[1], zones, priority)


def ids_from_xlsx(path: Path, sheet: str, start: int, col: int) -> tuple[list[str], int]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ids: list[str] = []
    rows = 0
    try:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=start, values_only=True):
            if len(row) > col and norm_id(row[col]):
                ids.append(norm_id(row[col]))
                rows += 1
    finally:
        wb.close()
    return ids, rows


def collect_fact_client_sets() -> list[dict[str, Any]]:
    sets: list[dict[str, Any]] = []
    for name, grain, path, sheet, start, col in FACT_SPECS:
        if not path.exists():
            continue
        try:
            ids, rows = ids_from_xlsx(path, sheet, start, col)
        except KeyError:
            continue
        sets.append({"source": name, "grain": grain, "rows": rows, "ids": ids})

    vogeo = WORK / "VENDEDORES PARTICULARES" / "UMAP" / "VOGEO.csv"
    if vogeo.exists():
        rows = read_csv_dicts(vogeo)
        sets.append({"source": "UMAP / VOGEO", "grain": "cliente volumen-importe", "rows": len(rows), "ids": [r.get("Cliente") for r in rows]})
    return sets


def fact_sources(clients: dict[str, ClientRecord]) -> list[dict[str, Any]]:
    sources: list[dict[str, Any]] = []
    geocoded = {cid for cid, client in clients.items() if client.coord}

    for fact_set in collect_fact_client_sets():
        unique = {norm_id(x) for x in fact_set["ids"] if norm_id(x)}
        hits = unique & geocoded
        sources.append(
            {
                "source": fact_set["source"],
                "grain": fact_set["grain"],
                "rows": fact_set["rows"],
                "unique_clients": len(unique),
                "geocoded_hits": len(hits),
                "coverage_pct": round(len(hits) / len(unique) * 100, 1) if unique else None,
            }
        )

    return sources


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)

    clients: dict[str, ClientRecord] = {}
    add_csv_clients(clients)
    add_clientes_totales(clients)
    add_volumen_clients(clients)
    assign_zones(clients)

    client_rows = []
    for client in sorted(clients.values(), key=lambda c: c.client_id):
        coord = client.coord
        client_rows.append(
            {
                "client_id": client.client_id,
                "name": client.name,
                "seller": client.seller,
                "status": client.status,
                "category": client.category,
                "address": client.address,
                "lat": coord[0] if coord else "",
                "lon": coord[1] if coord else "",
                "zone_id": client.zone_id,
                "zone_name": client.zone_name,
                "sources": "|".join(sorted(client.sources)),
            }
        )

    write_csv(
        OUT / "clients_normalized.csv",
        client_rows,
        ["client_id", "name", "seller", "status", "category", "address", "lat", "lon", "zone_id", "zone_name", "sources"],
    )

    zone_counts = Counter(row["zone_name"] for row in client_rows)
    zone_geo_counts = Counter(row["zone_name"] for row in client_rows if row["lat"] != "")
    zone_rows = [
        {"zone_name": zone, "clients": count, "geocoded_clients": zone_geo_counts.get(zone, 0)}
        for zone, count in sorted(zone_counts.items())
    ]
    write_csv(OUT / "zone_summary.csv", zone_rows, ["zone_name", "clients", "geocoded_clients"])

    seller_zone = Counter((row["seller"] or "(sin vendedor)", row["zone_name"]) for row in client_rows)
    seller_zone_rows = [
        {"seller": seller, "zone_name": zone, "clients": count}
        for (seller, zone), count in sorted(seller_zone.items())
    ]
    write_csv(OUT / "seller_zone_summary.csv", seller_zone_rows, ["seller", "zone_name", "clients"])

    conflict_rows = [row for row in client_rows if row["status"] == "CONFLICTO A/I"]
    write_csv(
        OUT / "status_conflicts.csv",
        conflict_rows,
        ["client_id", "name", "seller", "status", "category", "address", "lat", "lon", "zone_id", "zone_name", "sources"],
    )

    no_coordinate_rows = [row for row in client_rows if row["lat"] == ""]
    write_csv(
        OUT / "clients_without_coordinates.csv",
        no_coordinate_rows,
        ["client_id", "name", "seller", "status", "category", "address", "lat", "lon", "zone_id", "zone_name", "sources"],
    )

    fact_exception_rows = []
    for fact_set in collect_fact_client_sets():
        for client_id in sorted({norm_id(x) for x in fact_set["ids"] if norm_id(x)}):
            client = clients.get(client_id)
            if client and client.coord:
                continue
            fact_exception_rows.append(
                {
                    "source": fact_set["source"],
                    "grain": fact_set["grain"],
                    "client_id": client_id,
                    "in_client_master": "yes" if client else "no",
                    "has_coordinates": "yes" if client and client.coord else "no",
                    "name": client.name if client else "",
                    "seller": client.seller if client else "",
                    "status": client.status if client else "",
                }
            )
    write_csv(
        OUT / "fact_clients_without_map_point.csv",
        fact_exception_rows,
        ["source", "grain", "client_id", "in_client_master", "has_coordinates", "name", "seller", "status"],
    )

    fact_rows = fact_sources(clients)
    write_csv(OUT / "fact_source_coverage.csv", fact_rows, ["source", "grain", "rows", "unique_clients", "geocoded_hits", "coverage_pct"])

    total = len(client_rows)
    geocoded = sum(1 for row in client_rows if row["lat"] != "")
    no_coord = total - geocoded
    conflicts = len(conflict_rows)

    report = [
        "# Auditoria Comercial Embolsado",
        "",
        f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "## Resumen",
        "",
        f"- Clientes normalizados: {total}",
        f"- Clientes con coordenadas validas: {geocoded} ({geocoded / total * 100:.1f}%)" if total else "- Clientes con coordenadas validas: 0",
        f"- Clientes sin coordenadas validas: {no_coord}",
        f"- Clientes con conflicto activo/inactivo: {conflicts}",
        f"- Clientes/fuente analitica sin punto de mapa: {len(fact_exception_rows)}",
        "",
        "## Clientes por zona",
        "",
        "| Zona | Clientes | Geocodificados |",
        "|---|---:|---:|",
    ]
    for row in zone_rows:
        report.append(f"| {row['zone_name']} | {row['clients']} | {row['geocoded_clients']} |")

    report.extend(
        [
            "",
            "## Cobertura de fuentes analiticas",
            "",
            "| Fuente | Grano | Filas | Clientes unicos | Geocodificados | Cobertura |",
            "|---|---|---:|---:|---:|---:|",
        ]
    )
    for row in fact_rows:
        coverage = "" if row["coverage_pct"] is None else f"{row['coverage_pct']}%"
        report.append(
            f"| {row['source']} | {row['grain']} | {row['rows']} | {row['unique_clients']} | {row['geocoded_hits']} | {coverage} |"
        )

    report.extend(
        [
            "",
            "## Lectura operativa",
            "",
            "- El framework prioriza Excels actuales para ventas/actividad y usa coordenadas existentes como puente inicial.",
            "- Volumen por Cliente es la mejor primera capa para mapa por su alta cobertura geografica.",
            "- Actividad, ranking y tipo de producto son aptos para tablero y drilldown, pero requieren mejorar el puente de clientes.",
            "- Rentabilidad queda para una segunda integracion porque su grano principal es producto/planta/vendedor/dia, no cliente.",
            "",
            "## Archivos generados",
            "",
            "- clients_normalized.csv",
            "- zone_summary.csv",
            "- seller_zone_summary.csv",
            "- status_conflicts.csv",
            "- clients_without_coordinates.csv",
            "- fact_clients_without_map_point.csv",
            "- fact_source_coverage.csv",
        ]
    )
    REPORT.write_text("\n".join(report) + "\n", encoding="utf-8")
    print(f"Wrote {REPORT}")


if __name__ == "__main__":
    main()
